import openpyxl
import io


class Workbook(object):
	def __init__(self, filename, *args, **kwargs):
		self.full_path = filename
		self.workbook = self.__read_workbook(*args, **kwargs)

	def __read_workbook(self, *args, **kwargs):
		if self.full_path:
			default_kwargs = dict(read_only=True, data_only=True)
			try:
				with open(self.full_path, 'rb') as f:
					in_mem_file = io.BytesIO(f.read())
					return openpyxl.load_workbook(in_mem_file, *args, **{**default_kwargs, **kwargs})
			except PermissionError:
				print(f'Can not open {self.full_path}')
				return None
		return None

	@property
	def creator(self):
		return self.workbook.properties.creator

	@property
	def created(self):
		return self.workbook.properties.created

	@property
	def modified(self):
		return self.workbook.properties.modified

	@property
	def last_modified_by(self):
		return self.workbook.properties.lastModifiedBy

	@property
	def sheet_names(self):
		if self.workbook:
			return self.workbook.sheetnames
		return None

	@property
	def worksheets(self):
		if self.workbook:
			return self.workbook.worksheets
		return None

	@property
	def chartsheets(self):
		if self.workbook:
			return self.workbook.chartsheets
		return None

	def get_sheet_by_name(self, sheet_name):
		return self.workbook[sheet_name]

	def get_sheet_by_index(self, index):
		return self.workbook[self.workbook.sheetnames[index]]

	def get_range_by_name(self, sheet_name=None, range_address=None):
		if self.workbook:
			sheet_name = self.workbook[self.workbook.sheetnames[0]] if sheet_name is None else sheet_name
			ws = self.workbook[sheet_name]
			if range_address is None:
				return ws.get_squared_range(ws.min_column, ws.min_row, ws.max_column, ws.max_row)
			return ws[range_address]
		return None

	def get_range_by_square(self, sheet_name, min_row, min_column, max_row, max_column):
		if self.workbook:
			sheet_name = self.workbook[self.workbook.sheetnames[0]] if sheet_name is None else sheet_name
			return self.workbook[sheet_name].get_squared_range(min_column, min_row, max_column, max_row)
		return None
